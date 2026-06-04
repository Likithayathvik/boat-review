#Project Name : boAt - App Feedback Analyzer Tool
#Author       : boAt R&D (QA)
#Date         : 15 April 2024
#Version      : v2.3
#Description  : Google Console Play Store boAt App Feedback Analyzer
#Script & IDE : Python - Web Scraping [IDE Used - Thonny]
#********************************************************************#

#Library dependencies
#---------------------
import os
import re
import csv
import sys
import time
import uuid
import shutil
import smtplib
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from transformers import pipeline
from rapidfuzz import fuzz, process
import openpyxl
from openpyxl import load_workbook
from email.mime.text import MIMEText
from datetime import datetime, timedelta
from selenium.webdriver.common.by import By
from email.mime.multipart import MIMEMultipart
from selenium.webdriver.common.keys import Keys
from email.mime.application import MIMEApplication
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl.styles import Font, PatternFill, Alignment
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import ElementClickInterceptedException
import time
import undetected_chromedriver as uc

# Airdopes Atom and Airdopes Prime placed BEFORE Airdopes so more specific prefixes match first
DEVICE_TYPES = ["Stone", "Rockerz", "Airdopes Atom", "Airdopes Prime", "Airdopes", "Nirvana", "Party Pal", "immortal", "avante bar", "aavante", "party pal", "unity"]

# Alias map: normalise fuzzy-matched prefix → canonical display name
DEVICE_TYPE_ALIAS = {
    "aavante": "Aavante Bar",
    "avante bar": "Aavante Bar",
    "unity": "Boat Unity",
    "airpods":"Airdopes",
}

CATEGORY_KEYWORDS = ["battery information","charging information","Crash","battery", "Close","connect", "custom equalizer","Equalizer","Diagnostics",
                     "Login","log into","log", "Location", "tap", "button", "touch","EQ", "fw update","firmware update","Update",
                     "phone number","mobile number","otp",
                     "noise cancellation","anc",
                     "freez","privacy","policy","account","screen","download","load","stuck","support","warranty","app open","warranty","service","details",
                     "mydetails","personal information","personal","email address","email",
                     "control",
                     "compatible","customer care","complaint","ldac","tablet","tab","ipad",]

PRIORITY_KEYWORDS = [
    "custom equalizer",
    "tablet",
    "tab",
    "ipad",
    "log in",
    "firmware update",
    "fw update",
    "battery information",
    "charging information",
    "login",
    "app open",
    "phone number",
    "mobile number",
    "otp",
    "personal information",
    "email address",
    "mydetails",
    "connect",
    "connectivity",
    "sku not supported",
]

FORCE_UPDATE_SIGNALS = [
    "without updating", "without update", "let me use",
    "skip update", "force update", "forced to update",
    "dont want to update", "don't want to update",
    "not updating", "without upgrading"
]

THRESHOLD = 80
# Global variables declaration
global App_Rating
global App_Version
global lowest_date
global latest_date

products = [
    "Rockerz 480","Rockerz 421","Airdopes Ace Gen 2","Airdopes Alpha Gen 2",
    "Airdopes 213","Airdopes 212","Airdopes 219","Rockerz 113","Rockerz 112",
    "Rockerz 111","Airdopes Prime 513 ANC","Rockerz Zen ANC","Airdopes 131 Gen 2",
    "Rockerz Prime 415","Rockerz 301 ANC","Rockerz Plus 550 ANC","Rockerz 512 ANC",
    "Stone Arc Pro Plus","Stone Arc Pro","Aavante Bar 950","Aavante Bar Aspire",
    "Aavante Bar 490","Aavante Bar 485","Aavante Bar 480","Airdopes Prime 413",
    "Rockerz Trinity Grande","Airdopes 313","Rockerz Prime 255Z","Rockerz Prime 205",
    "Airdopes Prime 511","Airdopes Prime 412","Stone 358","Stone 352","Stone 350",
    "Stone 358 Pro","Stone 352 Pro","Stone 350 Pro","Airdopes 301","Airdopes 181 Pro",
    "Airdopes 161 ANC Elite","Rockerz Summit","Airdopes 91 Prime","Airdopes 141 Elite ANC",
    "Airdopes Ace","Airdopes 71","Airdopes Ultra Plus","Airdopes 131 Elite ANC",
    "Airdopes 101 v2","Airdopes Joy","Rockerz 235 Pro","Rockerz 430","Airdopes 121 Pro Plus",
    "Airdopes Plus 311","Airdopes 311 Pro","Airdopes 111v2","Rockerz Trinity Gen 2",
    "Airdopes Prime 701 ANC","Rockerz 450 Pro","Rockerz 450R","Rockerz 450",
    "Nirvana Ivy Pro","Nirvana Zenith Pro","Nirvana Iris","PartyPal 600","Rockerz 650 Pro",
    "Rockerz 551 ANC Pro","Nirvana X TWS","Airdopes 800 HiDef","Nirvana Crystl",
    "Airdopes Ultra Pro","Nirvana Lucid","Airdopes Alpha","Airdopes 161","Airdopes 141",
    "Airdopes 131","Nirvana Ion ANC Pro","Rockerz 210 ANC","Nirvana Ivy","Rockerz 255 Z Plus",
    "Stone Lumos","Nirvana Space","NIRVANA ZENITH","NIRVANA NEBULA","Nirvana Eutopia",
    "Airdopes Supreme","Rockerz 255 ANC","Nirvana Ion ANC","Nirvana Ion","Airdopes Flex 454 ANC",
    "Airdopes 800","Rockerz 255 Max","Airdopes 300","NIRVANA 525ANC","Airdopes 341ANC",
    "Airdopes 393ANC","Airdopes 172","Rockerz 255 Pro+","Rockerz 333 Pro","Rockerz 333",
    "Rockerz 330 Pro","Nirvana Crown","Airdopes Prime 700 ANC",
]

# Map Reason keywords to Consolidated Reason
CONSOLIDATION_MAP = {
    "custom equalizer": "Custom Equalizer issue",
    "equalizer": "Equalizer issue",
    "eq": "Equalizer issue",
    "option": "Button Customization",
    "tap": "Button Customization",
    "touch": "Button Customization",
    "button": "Button Customization",
    "customization": "Button Customization",
    "control": "Button Customization",
    "crash": "App Crash",
    "close": "App Crash",
    "app close": "App Crash",
    "connect": "Connectivity",
    "detect": "Connectivity",
    "fw update": "Firmware Update",
    "firmware update": "Firmware Update",
    "battery information": "Battery Information",
    "charging information": "Battery Information",
    "battery": "Battery Information",
    "noise cancellation": "Noise Cancellation Issue",
    "anc": "Noise Cancellation Issue",
    "login": "Login issue",
    "log into": "Login issue",
    "log": "Login issue",
    "issue": "App Issue",
    "app open": "App Issue",
    "phone number": "App Privacy Issue",
    "mobile number": "App Privacy Issue",
    "otp": "App Privacy Issue",
    "personal information": "App Privacy Issue",
    "email address": "App Privacy Issue",
    "email": "App Privacy Issue",
    "location": "Location Permission",
    "freez": "App Crash",
    "positive": "Positive Feedback",
    "update": "App Update issue",
    "app space": "App size",
    "download": "App download Issue",
    "privacy": "App Privacy Issue",
    "policy": "App Policy Issue",
    "account": "Login Issue",
    "screen": "App Issue",
    "load": "App Issue",
    "stuck": "App Issue",
    "support": "SKU not supported",
    "diagnostics": "Smart diagnostics",
    "warranty": "Service Issue",
    "service": "Service Issue",
    "details": "Privacy Issue",
    "mydetails": "App Privacy Issue",
    "personal": "App Privacy Issue",
    "compatible": "App not supported",
    "customer care": "Service Issue",
    "complaint": "Service Issue",
    "ldac": "App feature Issue",
    "tablet": "App not supported",
    "tab": "App not supported",
    "ipad": "App not supported",
}

from rapidfuzz import fuzz, process
import re
import unicodedata

#***********************************************************************#

def HearablesApp_WebScraping():
    # Create a Chrome instance using undetected-chromedriver
    options = uc.ChromeOptions()
    options.add_argument("--no-first-run --no-service-autorun --password-store=basic")

    # Start the driver
    driver = uc.Chrome(options=options, version_main=147)
    #driver = webdriver.Chrome()

    try:
        # Navigate directly to Play Console — Google will redirect to sign-in if needed
        driver.get('https://play.google.com/console/u/0/developers/7189426761488379489/app/4973043346552626474/user-feedback/reviews')
        time.sleep(10)

        # Find the email input field and enter your email
        email_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'identifierId'))
        )
        email_input.send_keys('boatqatest@gmail.com')
        email_input.send_keys(Keys.RETURN)

        time.sleep(5)

        # Wait for the password input field and enter password
        password_input = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.NAME, 'Passwd'))
        )
        password_input.send_keys('Testing@12345')
        password_input.send_keys(Keys.RETURN)

        time.sleep(10)

        try:
            continue_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[text()='Continue']"))
            )
            continue_button.click()
            print("Clicked Continue")
            time.sleep(5)
        except:
            print("No Continue button found or popup didn't appear")

        # 3️⃣ Handle pop-up and click "Cancel"
        try:
            cancel_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[text()='Cancel']"))
            )
            cancel_button.click()
            print("Cancelled popup")
            time.sleep(5)
        except:
            print("No Cancel button found or popup didn't appear")

        # Now, navigate to the desired page (in this case, Google Play Console)
        driver.get('https://play.google.com/console/u/0/developers/7189426761488379489/app/4973043346552626474/user-feedback/reviews')
        time.sleep(10)

        # Wait for the dropdown button to be clickable
        dropdown_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//div[@aria-label="Reviews per page 10 selected."]'))
        )

        # Click the dropdown button
        dropdown_button.click()
        print("Dropdown clicked")
        time.sleep(5)

        # IMPROVED: Use a more reliable selector for the "50" option
        try:
            # Wait for the dropdown options to appear
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'material-select-dropdown-item'))
            )

            # Find the "50" option by its text content
            option_50 = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//material-select-dropdown-item[.//span[text()='50']]"))
            )

            # Scroll into view if needed
            driver.execute_script("arguments[0].scrollIntoView(true);", option_50)
            time.sleep(1)

            # Try clicking
            try:
                option_50.click()
                print("Successfully clicked 50 items per page option")
            except ElementClickInterceptedException:
                # If regular click fails, use JavaScript click
                driver.execute_script("arguments[0].click();", option_50)
                print("Successfully clicked 50 items per page option using JavaScript")

            time.sleep(10)

            # Save the webpage to an HTML file
            with open("HearablesApp_1.html", "w", encoding="utf-8") as file:
                file.write(driver.page_source)
            print("Web page contents saved to 'HearablesApp_1.html'.")

            # Get the HTML content of the page after navigating to it
            html_content = driver.page_source

            # Parse the HTML using BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')

            # Define the CSS selector for the material-button with the specified aria-label
            css_selector = 'material-button[aria-label="Go to the next page"]'

            # Loop to click the element 11 times
            for i in range(11):
                # Wait for the element to be present on the page
                wait = WebDriverWait(driver, 10)
                element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, css_selector)))

                if element:
                    try:
                        # Check if the element is clickable
                        if element.is_enabled() and element.is_displayed():
                            # Click the element
                            element.click()
                            print(f"Clicked the material-button with aria-label 'Go to the next page' (Click {i + 1}/12).")
                            time.sleep(10)
                            # Save the page source to an HTML file
                            with open(f'HearablesApp_{i+2}.html', 'w', encoding='utf-8') as file:
                                file.write(driver.page_source)
                            print(f"Page details saved in 'HearablesApp_{i+2}.html'.")
                        else:
                            print("The material-button is not clickable.")
                    except ElementClickInterceptedException as e:
                        print("Element click intercepted: ", e)
                else:
                    print("The material-button with aria-label 'Go to the next page' is not available on the page.")

        except Exception as e:
            print(f"Error selecting 50 items per page: {e}")
            print("Trying alternative method...")

            # Alternative method - find all items and click the one with text "50"
            try:
                all_items = driver.find_elements(By.CSS_SELECTOR, 'material-select-dropdown-item span.label')

                for item in all_items:
                    if item.text.strip() == '50':
                        try:
                            item.click()
                            print("Clicked on 50 using alternative method")
                            time.sleep(10)
                            break
                        except:
                            driver.execute_script("arguments[0].click();", item)
                            print("Clicked on 50 using JavaScript (alternative method)")
                            time.sleep(10)
                            break
            except Exception as e2:
                print(f"Alternative method also failed: {e2}")

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Close the WebDriver
        driver.quit()

def find_elements_in_review_container(html_content):

    global App_Rating_Value

    soup = BeautifulSoup(html_content, 'html.parser')
    review_containers = soup.find_all(class_='review-container')

    data = []

    for container in review_containers:
        App_Rating = container.find(attrs={'debug-id': 'google-play-rating'})
        App_Rating_Value = App_Rating.find('span', class_='value').text.strip() if App_Rating else "NA"

        author_display_name = container.find(class_='author-display-name')
        last_update_time = container.find(class_='last-update-time')
        device_display_name = container.find(attrs={'debug-id': 'device-display-name'})
        version_code = container.find(attrs={'debug-id': 'version-code'})
        version_name = container.find(attrs={'debug-id': 'version-name'})
        os_version = container.find(attrs={'debug-id': 'os-version'})
        review_body = container.find(attrs={'debug-id': 'review-body'})

        star_icons = container.find(attrs={'debug-id': 'star-icons'})
        star_icons_value = star_icons.get('aria-label').split(' ')[0] if star_icons else "NA"

        conversation_content = container.find(class_='conversation-content')

        # If any element is not found, set its value to "NA"
        author_display_name_value = author_display_name.text.strip() if author_display_name else "NA"
        last_update_time_value = last_update_time.text.strip() if last_update_time else "NA"
        device_display_name_value = device_display_name.text.strip() if device_display_name else "NA"
        version_code_value = version_code.text.strip() if version_code else "NA"
        version_name_value = version_name.text.strip() if version_name else "NA"
        os_version_value = os_version.text.strip() if os_version else "NA"
        review_body_value = review_body.text.strip() if review_body else "NA"
        conversation_content_value = conversation_content.text.strip() if conversation_content else "NA"

        # Append data to the list
        data.append([
            author_display_name_value,
            star_icons_value,
            review_body_value,
            device_display_name_value,
            os_version_value,
            last_update_time_value,
            version_code_value,
            version_name_value,
            conversation_content_value
        ])

    return data

def All_HtmlFiles_Csv():
    directory_path = os.path.dirname(os.path.realpath(__file__))
    folder_path = directory_path  # Replace with the path to your folder

    # Get a list of all files in the folder
    file_list = os.listdir(folder_path)

    # Filter the list to include only HTML files
    html_files = [file for file in file_list if file.endswith('.html')]

    # List to store data from all HTML files
    all_data = []

    # Process each HTML file
    for html_file in html_files:
        # Wait for 10 seconds before processing the next file
        time.sleep(10)
        file_path = os.path.join(folder_path, html_file)
        with open(file_path, 'r', encoding='utf-8') as file:
            html_content = file.read()
        print(f"Processing: {file_path}")

        # Get data for the current HTML file
        data = find_elements_in_review_container(html_content)

        # Append data to the list for all HTML files
        all_data.extend(data)

    # Write data to a single CSV file
    csv_file_path = os.path.join(directory_path, 'Hearables_Data.csv')
    with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
        csv_writer = csv.writer(csvfile)

        # Write header
        header = ['User Name', 'Ratings', 'Reviews',
                  'Phone Name', 'OS Version', 'Date Time',
                  'Version Code', 'App Version', 'boAt Reply']
        csv_writer.writerow(header)

        # Write data rows
        csv_writer.writerows(all_data)

def Csv_Date_Rating_Filter_Alldate():
    # Load the CSV file into a DataFrame
    df = pd.read_csv('Hearables_Data.csv')

    # Convert the "Date Time" column to a datetime data type if it's not already
    df['Date Time'] = pd.to_datetime(df['Date Time'])
    df['Date Time'] = pd.to_datetime(df['Date Time'], format="%b %d, %Y, %H:%M")

    # Remove the filtering for the last 7 days
    # Comment out or remove the following two lines
    # seven_days_ago = datetime.now() - timedelta(days=7)
    # df = df[df['Date Time'] >= seven_days_ago]

    # Sort the DataFrame in descending order based on the "Date Time" column
    df = df.sort_values(by='Date Time', ascending=False)

    # Filter the "Ratings" column to keep only values 1, 2, and 3
    df = df[df['Ratings'].isin([1, 2, 3, 4, 5])]

    # Save the filtered DataFrame to a new CSV file
    df.to_csv('HearablesAppGCWeb.csv', index=False)

    time.sleep(2)

    # Load the CSV file into a pandas DataFrame
    df = pd.read_csv('HearablesAppGCWeb.csv')

    # Create a new column 'reviewId' with unique identifiers
    df['reviewId'] = [str(uuid.uuid4()) for _ in range(len(df))]

    # Reorder the columns to place 'reviewId' in the 'A' column
    cols = df.columns.tolist()
    cols = ['reviewId'] + [col for col in cols if col != 'reviewId']
    df = df[cols]

    # Save the DataFrame with 'reviewId' column to the same CSV file
    df.to_csv('HearablesAppGCWeb.csv', index=False)

    print("Date and Ratings filter is successfully completed and saved into HearablesAppGCWeb CSV file")


def Csv_Date_Rating_Filter_7days():
    # Load the CSV file into a DataFrame
    df = pd.read_csv('Hearables_Data.csv')
    # Convert the "Date Time" column to a datetime data type if it's not already
    #df['Date Time'] = pd.to_datetime(df['Date Time'])
    df['Date Time'] = pd.to_datetime(df['Date Time'], format="%b %d, %Y, %H:%M")
    # Calculate the date 7 days ago from the current date
    seven_days_ago = datetime.now() - timedelta(days=7)
    # Filter the DataFrame to keep only rows with dates from the last 7 days
    filtered_df = df[df['Date Time'] >= seven_days_ago]
    # Sort the filtered DataFrame in descending order based on the "Date Time" column
    #filtered_df = filtered_df.sort_values(by='Date Time', ascending=False)
    # Filter the "Ratings" column to keep only values 1 to 5
    filtered_df = filtered_df[filtered_df['Ratings'].isin([1, 2, 3, 4, 5])]
    # Save the filtered DataFrame to a new CSV file
    filtered_df.to_csv('HearablesAppGCWeb.csv', index=False)
    time.sleep(2)
    # Load the CSV file into a pandas DataFrame
    df = pd.read_csv('HearablesAppGCWeb.csv')
    # Create a new column 'reviewId' with unique identifiers
    df['reviewId'] = [str(uuid.uuid4()) for _ in range(len(df))]
    # Reorder the columns to place 'reviewId' in the 'A' column
    cols = df.columns.tolist()
    cols = ['reviewId'] + [col for col in cols if col != 'reviewId']
    df = df[cols]
    df.to_csv('HearablesAppGCWeb.csv', index=False)

    print("Date and Ratings filter is successfully completed and saved into HearablesAppGCWeb CSV file")

def CSV_Remove_Duplicates():
    # Get the current directory path
    directory_path = os.path.dirname(os.path.realpath(__file__))

    # Specify the folder path where the CSV files are located
    folder_path = directory_path  # Replace with the path to your folder

    # Check if the folder path exists
    if os.path.exists(folder_path):
        # List all files in the folder
        files = os.listdir(folder_path)

        # Loop through each file in the folder
        for file_name in files:
            # Check if the file is a CSV file
            if file_name.endswith('.csv'):
                # Construct the full path to the CSV file
                file_path = os.path.join(folder_path, file_name)

                # Read the CSV file into a pandas DataFrame
                df = pd.read_csv(file_path)

                # Count and remove duplicates based on the specified columns
                duplicate_count_before = df.duplicated(subset=['User Name', 'Ratings', 'Reviews', 'Phone Name']).sum()
                df = df.drop_duplicates(subset=['User Name', 'Ratings', 'Reviews', 'Phone Name'])

                # Overwrite the original CSV file with the cleaned DataFrame
                df.to_csv(file_path, index=False)

                # Print the number of duplicate rows removed for each file
                print(f'For {file_name}: {duplicate_count_before} duplicate rows removed.')
    else:
        print(f'The specified folder path "{folder_path}" does not exist.')

def Csv_2_Xlsx():

    directory_path = os.path.dirname(os.path.realpath(__file__))  # Get the current script's directory
    csv_file = 'HearablesAppGCWeb.csv'
    xlsx_file = 'Android_HearablesApp_Review.xlsx'
    # Read the CSV file using pandas.
    df = pd.read_csv(csv_file)

    # Write the DataFrame to an XLSX file.
    df.to_excel(xlsx_file, index=False)

    print(f'CSV file "{csv_file}" has been converted to XLSX file "{xlsx_file}".')

def Positive_Reviews():
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('Android_HearablesApp_Review.xlsx')

    # Filter the DataFrame to keep only rows with ratings 4 and 5
    positive_reviews_df = df[df['Ratings'].isin([4, 5])]

    # Save the positive reviews to the same Excel file in a new sheet named 'Positive_Reviews'
    with pd.ExcelWriter('Android_HearablesApp_Review.xlsx', engine='openpyxl', mode='a') as writer:
        positive_reviews_df.to_excel(writer, sheet_name='Positive_Reviews', index=False)

    print("Positive reviews saved to 'Android_HearablesApp_Review.xlsx' under sheet name 'Positive_Reviews'.")

def normalize_unicode_text(text):
    """Convert stylized Unicode characters (𝓷𝓲𝓬𝓮 → nice) to plain ASCII."""
    return unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')


def extract_device_names(text, threshold=70):
    """
    Device extraction with:
    ✔ Fuzzy match only for prefix (Device Type)
    ✔ Keep numeric model EXACT — picks first number AFTER prefix and BEFORE suffix
    ✔ EXACT suffix detection
    ✔ Prefix + alpha words (no number): e.g. "Stone Spinx Pro", "Nirvana Lucid"
    ✔ Aavante-specific: captures Prefix + version (e.g. 2.0) + model number
      e.g. "Boat Aavante 2.0 150" → "Aavante Bar 2.0 150"
      Handles "aavante" (double-a) spelling via DEVICE_TYPE_ALIAS.
    """
    if not text or not isinstance(text, str):
        return "Unknown Device"
    if len(text.split()) <= 2:
        return "Unknown Device"

    txt = text.lower()

    # ── Capture version strings (e.g. "2.0") BEFORE tokenising ──────────────
    version_strings = re.findall(r'\b\d+\.\d+\b', txt)   # e.g. ["2.0"]

    # 1️⃣ Fuzzy match PREFIX only
    words = re.findall(r"[a-zA-Z0-9]+", txt)
    windows = words + [" ".join(words[i:i+2]) for i in range(len(words)-1)]
    best_prefix = None
    best_score = 0
    for w in windows:
        result = process.extractOne(
            w, [p.lower() for p in DEVICE_TYPES], scorer=fuzz.ratio
        )
        if result and result[1] >= threshold and (
            result[1] > best_score or
            (result[1] == best_score and best_prefix and len(result[0]) > len(best_prefix))
        ):
            best_prefix = result[0]
            best_score = result[1]

    # Resolve alias → canonical display name (e.g. "aavante" → "Aavante Bar")
    canonical_prefix = DEVICE_TYPE_ALIAS.get(best_prefix, best_prefix.title() if best_prefix else None)

    # 2️⃣ Strip version strings, then find all numbers with their positions
    txt_no_version = re.sub(r'\b\d+\.\d+\b', '', txt)
    nums = re.findall(r"\d+", txt_no_version)

    SUFFIX_MAP = {
        "anc elite": "ANC Elite",
        "anc pro": "ANC Pro",
        "ancpro": "ANC Pro",
        "anc": "ANC",
        "pro+": "Pro+",
        "pro plus": "Pro+",
        "pro": "Pro",
        "plus": "Plus",
        "prime": "Prime",
        "gen 2": "Gen 2",
        "gen2": "Gen 2",
        "elite": "Elite",
        "ultra pro": "Ultra Pro",
        "ultra": "Ultra",
        "max": "Max",
        "v1": "v1",
        "v2": "v2",
        "v2 pro": "V2 Pro",
        "v2pro": "V2 Pro",
        "r": "R",
        "z plus": "Z Plus",
        "zplus": "Z Plus",
        "z": "Z",
        "TWS": "TWS",
        "x tws": "X TWS",
        "xtws": "X TWS",
        "hidef": "HiDef",
        "hi def": "HiDef",
        "ionanc": "Ion ANC",
        "ion anc": "Ion ANC",
        "ion anc pro": "Ion Anc Pro",
        "ionancpro": "Ion Anc Pro",
        "ion": "ion",
        "loop": "loop",
        "tws": "TWS",
        "v2Pro": "V2 Pro",
        "v2 Pro": "V2 Pro",
        "crown": "Crown",
        "ivy": "Ivy",
        "IvyPro": "IvyPro",
        "Ivy Pro": "Ivy Pro",
    }

    # ── Pick model_num: first number that appears AFTER the prefix
    #    AND BEFORE the earliest suffix match in the text ──────────────────────
    model_num = None
    if nums and best_prefix:
        # Find where the prefix ends in txt_no_version
        prefix_first_word = best_prefix.split()[0]
        prefix_match = re.search(re.escape(prefix_first_word), txt_no_version)
        prefix_end_pos = prefix_match.end() if prefix_match else 0

        # Find where the earliest suffix begins in txt_no_version
        suffix_start_pos = len(txt_no_version)  # default: end of string
        for key in sorted(SUFFIX_MAP.keys(), key=lambda x: -len(x)):
            m = re.search(r"\b" + re.escape(key) + r"\b", txt_no_version)
            if m and m.start() > prefix_end_pos:
                if m.start() < suffix_start_pos:
                    suffix_start_pos = m.start()

        # All numbers with positions
        num_positions = [(m.group(), m.start()) for m in re.finditer(r"\d+", txt_no_version)]

        # Pick the first number that falls strictly between prefix_end and suffix_start
        between = [(n, pos) for n, pos in num_positions
                   if pos >= prefix_end_pos and pos < suffix_start_pos]
        if between:
            model_num = between[0][0]
        else:
            # Fallback: first number after prefix (ignore suffix boundary)
            if suffix_start_pos == len(txt_no_version):
                after_prefix = [(n, pos) for n, pos in num_positions if pos >= prefix_end_pos]
                if after_prefix:
                    model_num = after_prefix[0][0]
            # No number after prefix at all → model_num stays None
    elif nums:
        # No prefix found — keep old largest-number fallback
        model_num = sorted(nums, key=lambda x: (len(x), int(x)), reverse=True)[0]

    # ── Aavante-specific: Prefix + version + model number ────────────────────
    if best_prefix and "avante" in best_prefix.lower() and version_strings and model_num:
        version_str = version_strings[0]   # e.g. "2.0"
        detected_suffix = ""
        for key in sorted(SUFFIX_MAP.keys(), key=lambda x: -len(x)):
            if re.search(r"\b" + re.escape(key) + r"\b", txt):
                detected_suffix = SUFFIX_MAP[key]
                break
        device = f"{canonical_prefix} {version_str} {model_num}"
        if detected_suffix:
            device += f" {detected_suffix}"
        return device.strip()

    # -------------------------
    #  Prefix + alpha + optional suffix (no number in text)
    # -------------------------
    if best_prefix and not model_num:
        prefix_tokens = best_prefix.split()
        prefix_end = -1

        # Locate where the prefix ends in the word list
        for i in range(len(words)):
            window = " ".join(words[i:i + len(prefix_tokens)])
            if fuzz.ratio(window, best_prefix) >= threshold:
                prefix_end = i + len(prefix_tokens)
                break

        if prefix_end != -1:
            # --- detect the longest matching suffix present in text ---
            detected_suffix = ""
            for key in sorted(SUFFIX_MAP.keys(), key=lambda x: -len(x)):
                if re.search(r"\b" + re.escape(key) + r"\b", txt):
                    detected_suffix = SUFFIX_MAP[key]
                    break

            suffix_tokens = detected_suffix.lower().split() if detected_suffix else []
            middle_parts = []
            for w in words[prefix_end:]:
                # stop if we've reached the suffix word(s)
                if suffix_tokens and w == suffix_tokens[0]:
                    break
                # collect only alphabetic words (skip noise / numbers)
                if re.match(r'^[a-zA-Z]{2,}$', w):
                    middle_parts.append(w)
                    # no suffix expected → one model-name word is enough, stop
                    if not detected_suffix:
                        break
                else:
                    break   # stop at any digit/punctuation token

            parts = [canonical_prefix]
            if middle_parts:
                parts.append(" ".join(w.title() for w in middle_parts))
            if detected_suffix:
                parts.append(detected_suffix)

            result_device = " ".join(parts).strip()
            return result_device

        # Alpha-only fallback: fuzzy match full review against products list
        fallback = process.extractOne(
            txt, [p.lower() for p in products], scorer=fuzz.partial_ratio
        )
        if fallback and fallback[1] >= THRESHOLD:
            matched_idx = [p.lower() for p in products].index(fallback[0])
            return products[matched_idx]
        return "Unknown Device"

    if not best_prefix:
        detected_suffix = ""
        for key in sorted(SUFFIX_MAP.keys(), key=lambda x: -len(x)):
            for num in nums:
                pattern = rf"{num}\s*{re.escape(key)}\b"
                if re.search(pattern, txt, flags=re.IGNORECASE):
                    detected_suffix = SUFFIX_MAP[key]
                    break
            if detected_suffix:
                break
        if not detected_suffix:
            for key in sorted(SUFFIX_MAP.keys(), key=lambda x: -len(x)):
                for num in nums:
                    pattern = rf"{num}{re.escape(key)}\b"
                    if re.search(pattern, txt, flags=re.IGNORECASE):
                        detected_suffix = SUFFIX_MAP[key]
                        break
                if detected_suffix:
                    break
        if detected_suffix:
            return f"{model_num} {detected_suffix}".strip()
        # No-prefix fallback: fuzzy match full review against products list
        fallback = process.extractOne(
            txt, [p.lower() for p in products], scorer=fuzz.partial_ratio
        )
        if fallback and fallback[1] >= THRESHOLD:
            matched_idx = [p.lower() for p in products].index(fallback[0])
            return products[matched_idx]
        return "Unknown Device"

    detected_suffix = ""
    for key in sorted(SUFFIX_MAP.keys(), key=lambda x: -len(x)):
        if re.search(r"\b" + re.escape(key) + r"\b", txt):
            detected_suffix = SUFFIX_MAP[key]
            break
    if not detected_suffix:
        for num in nums:
            for key in sorted(SUFFIX_MAP.keys(), key=lambda x: -len(x)):
                pattern = rf"{num}{re.escape(key)}\b"
                if re.search(pattern, txt):
                    detected_suffix = SUFFIX_MAP[key]
                    break
            if detected_suffix:
                break

    # 4️⃣ Build device
    device = f"{canonical_prefix} {model_num}" if canonical_prefix else "Unknown Device"
    if detected_suffix:
        device += f" {detected_suffix}"

    # 5️⃣ FINAL FALLBACK: if still Unknown Device, fuzzy-match the full review
    #    text directly against the products list.
    if device.strip() == "Unknown Device":
        result = process.extractOne(
            txt, [p.lower() for p in products], scorer=fuzz.partial_ratio
        )
        if result and result[1] >= THRESHOLD:
            matched_idx = [p.lower() for p in products].index(result[0])
            device = products[matched_idx]

    return device.strip()


# Remove emojis before sentiment analysis
def remove_emojis(text):
    emoji_pattern = re.compile(
        "["
        "\U0001F600-\U0001F64F"  # emoticons
        "\U0001F300-\U0001F5FF"  # symbols & pictographs
        "\U0001F680-\U0001F6FF"  # transport & map symbols
        "\U0001F1E0-\U0001F1FF"  # flags (iOS)
        "\U00002700-\U000027BF"  # dingbats
        "\U0001F900-\U0001F9FF"  # supplemental symbols
        "\U00002600-\U000026FF"  # miscellaneous symbols
        "]+",
        flags=re.UNICODE,
    )
    return emoji_pattern.sub("", text)


EXCEPTION_PHRASES = ["even though", "even", "though","issue"]
def contains_exception_phrase(text):
    """
    Returns True if any exception phrase matches fuzzily above threshold.
    """
    text_lower = text.lower()
    for phrase in EXCEPTION_PHRASES:
        # Sliding window over words to check fuzzy match
        words = text_lower.split()
        for i in range(len(words)):
            # check up to len(phrase.split())+1 words together
            window = " ".join(words[i:i+len(phrase.split())+1])
            score = fuzz.partial_ratio(phrase, window)
            if score >= THRESHOLD:
                return True
    return False

def is_valid_device(d, products_set):
    """
    Check if device d is valid.
    - If numeric: match only against numeric parts of products_set
    - Otherwise: exact match in products_set
    """
    if d.isdigit():
        for product in products_set:
            # Extract all numeric parts from product name
            nums = re.findall(r'\d+', product)
            if d in nums:
                return True
        return False
    else:
        return d in products_set

def is_sentiment_positive(review, sentiment_analyzer):
    if not isinstance(review, str):
        return False
    cleaned = remove_emojis(review).strip()
    if not cleaned or len(cleaned.split()) > 5:
        return False
    try:
        return sentiment_analyzer(cleaned)[0]['label'].lower() == 'positive'
    except Exception:
        return False


def best_keyword_match(review_lower, keywords, threshold):
    """
    Three-tier matching:
      Tier 1 (priority exact): keyword is in PRIORITY_KEYWORDS AND is a literal
                               substring of the review. First match in
                               PRIORITY_KEYWORDS order wins immediately.
                               Also fuzzy-matches multi-word priority keywords
                               for typo resilience.
      Tier 2 (exact substring): keyword appears literally but is NOT a priority kw.
                               Among all tier-2 hits, longest keyword wins.
      Tier 3 (fuzzy only):     no exact match found at all; highest partial_ratio
                               wins (ties broken by keyword length).
    Returns "Other" if no keyword reaches `threshold`.
    """
    # Tier 1: priority keywords — exact first, then fuzzy for multi-word (typo resilience)
    for pkw in PRIORITY_KEYWORDS:
        pkw_lower = pkw.lower()
        if pkw_lower in review_lower:
            return pkw_lower
        if len(pkw_lower.split()) > 1:
            score = fuzz.partial_ratio(pkw_lower, review_lower)
            if score >= threshold:
                return pkw_lower

    # Tier 2: exact substring, non-priority — longest wins
    exact_kw  = None
    exact_len = 0

    # Tier 3: fuzzy fallback
    fuzzy_kw    = "Other"
    fuzzy_score = 0
    fuzzy_len   = 0

    for kw in keywords:
        kw_lower = kw.lower()
        if kw_lower in review_lower:
            if len(kw_lower) > exact_len:
                exact_kw  = kw_lower
                exact_len = len(kw_lower)
            continue

        score = fuzz.partial_ratio(kw_lower, review_lower)
        if score < threshold:
            continue
        if score > fuzzy_score or (score == fuzzy_score and len(kw_lower) > fuzzy_len):
            fuzzy_kw    = kw_lower
            fuzzy_score = score
            fuzzy_len   = len(kw_lower)

    return exact_kw if exact_kw is not None else fuzzy_kw


def Negative_Reviews():
    df = pd.read_excel('Android_HearablesApp_Review.xlsx')

    # ── Normalize stylized Unicode in Reviews column so readable text is stored ──
    df['Reviews'] = df['Reviews'].apply(
        lambda x: normalize_unicode_text(str(x)).strip() if pd.notna(x) else x
    )

    negative_reviews_df = df[df['Ratings'].isin([1,2,3])].copy()
    sentiment_analyzer = pipeline(
    "sentiment-analysis",
    model="distilbert/distilbert-base-uncased-finetuned-sst-2-english",
    cache_dir="./model_cache"
)

    reasons, DUT = [], []
    for review in negative_reviews_df['Reviews']:
        # ── Empty / NaN ──────────────────────────────────────────────────────
        if pd.isna(review) or not str(review).strip():
            reasons.append("Generic Comments")
            DUT.append("Unknown Device")
            continue

        if is_sentiment_positive(review, sentiment_analyzer):
            reasons.append("positive")
            DUT.append(extract_device_names(review))
            continue

        review_lower = review.lower()
        matched_category = best_keyword_match(review_lower, CATEGORY_KEYWORDS, THRESHOLD)

        # ── Fallback to full sentiment if no keyword matched ──────────────────
        if matched_category == "Other":
            clean_review = remove_emojis(review)
            if contains_exception_phrase(clean_review):
                sentiment = "neutral"
            else:
                sentiment = sentiment_analyzer(clean_review)[0]['label'].lower()
            if sentiment == "positive":
                matched_category = "positive"

        reasons.append(matched_category)
        DUT.append(extract_device_names(review))

    negative_reviews_df['Reason'] = reasons
    negative_reviews_df['DUT'] = DUT

    # ── Consolidation ─────────────────────────────────────────────────────────
    consolidated_reasons = []
    products_set = set(products)
    reviews_list = negative_reviews_df['Reviews'].tolist()

    for review, reason, dut in zip(reviews_list, reasons, DUT):
        r = reason.lower()
        d = dut.strip()
        review_lower = str(review).lower() if pd.notna(review) else ""

        # ── SKU not supported: device extracted but not in known products list ──
        if d != "Unknown Device" and not is_valid_device(d, products_set):
            consolidated = "SKU not supported"

        # ── Force update vs regular update issue ──────────────────────────────
        elif r == "update":
            if any(sig in review_lower for sig in FORCE_UPDATE_SIGNALS):
                consolidated = "Force App Update Not Required"
            else:
                consolidated = "App Update issue"

        # ── Standard consolidation map ─────────────────────────────────────────
        else:
            consolidated = CONSOLIDATION_MAP.get(r, "Generic Comments")

        consolidated_reasons.append(consolidated)

    negative_reviews_df['Consolidated Reason'] = consolidated_reasons

    with pd.ExcelWriter('Android_HearablesApp_Review.xlsx', engine='openpyxl', mode='a') as writer:
        negative_reviews_df.to_excel(writer, sheet_name='Negative_Reviews', index=False)

    print("Negative reviews and reasons saved to 'Negative_Reviews' sheet.")

def SheetNameChange():
    directory_path = os.path.dirname(os.path.realpath(__file__))  # Get the current script's directory
    # Load the Excel file
    file_path = 'Android_HearablesApp_Review.xlsx'
    current_sheet_name = 'Sheet1'
    new_sheet_name = 'Happ_Review_Data'

    try:
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)

        # Check if the current sheet name exists
        if current_sheet_name in wb.sheetnames:
            # Get the sheet object
            sheet = wb[current_sheet_name]

            # Change the sheet name to 'Negative_Reviews'
            sheet.title = new_sheet_name

            # Save the modified workbook with the new sheet name
            wb.save(file_path)

            print(f"The sheet name '{current_sheet_name}' has been changed to '{new_sheet_name}'.")
        else:
            print(f"The sheet '{current_sheet_name}' does not exist in the Excel file.")

        # Close the Excel file
        wb.close()

    except FileNotFoundError:
        print(f"The file '{file_path}' does not exist.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")


def Sorting_Xlsx(sheet_name):
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('Android_HearablesApp_Review.xlsx', sheet_name=sheet_name)

    # Load the XLSX file
    workbook = openpyxl.load_workbook('Android_HearablesApp_Review.xlsx')

    # Select the worksheet where you want to set column widths
    worksheet = workbook[sheet_name]

    # Define the column widths you want to set (e.g., columns A, B, and C)
    column_widths = {'A': 15, 'B': 15, 'C': 8, 'D': 42, 'E': 12.5, 'F': 11, 'G': 11, 'H': 10, 'I': 10, 'J': 48, 'K': 15, 'L': 21} # You can adjust the values as needed

    # Set the column widths
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    # Iterate through the rows and columns to apply wrap text and middle align
    for row in worksheet.iter_rows():
        for cell in row:
            # Set wrap text to True
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="center")

    # Save the workbook with changes to the same file
    workbook.save('Android_HearablesApp_Review.xlsx')

    # Close the workbook
    workbook.close()

    print(f"Xlsx modification done for sheet '{sheet_name}'.")

def Sorting_Summary_Xlsx(sheet_name):
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('Android_HearablesApp_Review.xlsx', sheet_name=sheet_name)

    # Load the XLSX file
    workbook = openpyxl.load_workbook('Android_HearablesApp_Review.xlsx')

    # Select the worksheet where you want to set column widths
    worksheet = workbook[sheet_name]

    # Define the column widths you want to set (e.g., columns A, B, and C)
    column_widths = {'A': 25, 'B': 15, 'C': 21, 'D': 23, 'E': 25, 'F': 28} # You can adjust the values as needed

    # Set the column widths
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    # Iterate through the rows and columns to apply wrap text and middle align
    for row in worksheet.iter_rows():
        for cell in row:
            # Set wrap text to True
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="center")

    # Save the workbook with changes to the same file
    workbook.save('Android_HearablesApp_Review.xlsx')

    # Close the workbook
    workbook.close()

    print(f"Xlsx modification done for sheet '{sheet_name}'.")


def Remove_Html():

    directory_path = os.path.dirname(os.path.realpath(__file__))

    folder_path = directory_path  # Replace with the path to your folder

    # Check if the folder path exists
    if os.path.exists(folder_path):
        # List all files in the folder
        files = os.listdir(folder_path)

        # Iterate through the files and delete HTML files
        for file in files:
            if file.endswith(".html"):
                file_path = os.path.join(folder_path, file)
                os.remove(file_path)
                print(f"Deleted: {file_path}")
    else:
        print(f"The folder path '{folder_path}' does not exist.")

def AppRating_find_element_in_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()

    # Parse HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all div elements with class "value-and-description"
    div_elements = soup.find_all('div', class_='value-and-description')

    # Check if any span element with class "value" is found within the div elements
    for div_element in div_elements:
        span_elements = div_element.find_all('span', class_='value')
        if span_elements:
            return span_elements[0].get_text()

    return None

def App_Rating_Finder():
    global App_Rating

    # Path to your HTML file
    html_file_path = 'HearablesApp_1.html'

    # Check if file exists before trying to open it
    if not os.path.exists(html_file_path):
        print(f"ERROR: {html_file_path} not found!")
        App_Rating = "N/A"
        return

    # Check if the element is available in the HTML file
    found_element = AppRating_find_element_in_html(html_file_path)

    if found_element:
        App_Rating = found_element[:3]
        print("App Rating Value found in the HTML file:", App_Rating)
    else:
        print("App Rating Value not found in HTML file")
        App_Rating = "N/A"

def App_Version_Finder():
    global App_Version

    # Load the Excel file
    workbook = openpyxl.load_workbook('Android_HearablesApp_Review.xlsx')

    # Assuming the data is in the first sheet, you can change the sheet name if it's different
    sheet = workbook.active

    # Assuming "App Version" is in column A, change the column index if it's different
    App_Version_column = 9

    # Assuming the first row contains headers and data starts from the second row
    data_start_row = 2

    # Initialize a variable to store the highest app version
    highest_version = None

    # Iterate through each row after the header row
    for row in range(data_start_row, sheet.max_row + 1):
        App_Version = sheet.cell(row=row, column=App_Version_column).value
        if App_Version:
            if highest_version is None or App_Version > highest_version:
                highest_version = App_Version

    highest_version_info = highest_version.split(":")[-1].strip()

    App_Version = highest_version_info

    # Print the highest app version
    print("Highest App Version:", App_Version)

def Review_Date_Calculator():
    global lowest_date
    global latest_date
    # Load the Excel file
    df = pd.read_excel('Android_HearablesApp_Review.xlsx')

    # Replace 'DateTimeColumn' with the actual name of the column containing your date and time data
    df['Date Time'] = pd.to_datetime(df['Date Time'])

    # Extract date part
    df['DateOnly'] = df['Date Time'].dt.date

    # Find the earliest and latest date
    lowest_date = df['DateOnly'].min()
    latest_date = df['DateOnly'].max()


    # Print the results
    print("Lowest Date in the DateTimeColumn:", lowest_date)
    print("Latest Date in the DateTimeColumn:", latest_date)


def Negative_Review_Analysis():
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel('Android_HearablesApp_Review.xlsx', sheet_name='Negative_Reviews')

    # Count occurrences of each "Consolidated Reason"
    reason_counts = df['Consolidated Reason'].value_counts()

    # Print the counts for each "Consolidated Reason"
    print("Negative Review Analysis:")
    for reason, count in reason_counts.items():
        print(f"{reason}: {count}")

    return reason_counts

def Summary():
    global App_Rating
    global App_Version
    global lowest_date
    global latest_date
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
    import math

    print("Generating styled Summary1 sheet...")

    # Load main review data
    df = pd.read_excel('Android_HearablesApp_Review.xlsx')

    # Review counts
    pos_count = len(df[df['Ratings'].isin([4, 5])])
    neg_count = len(df[df['Ratings'].isin([1, 2, 3])])

     # Calculate summary statistics
    app_name = "Android Hearables App " + str(App_Version)	# Replace with your app name
    app_rating = App_Rating + '*'
    review_data = str(lowest_date) + ' to ' + str(latest_date)

    # Define app details data
    app_details_data = {
        'App Name': [app_name],
		'Latest App Version': [App_Version],  # Keep the app version unchanged
        'Review Date': [review_data]
    }
    app_df = pd.DataFrame(app_details_data)
    counts_df = pd.DataFrame({
        'Category': ['Positive Reviews (4 & 5)', 'Negative Reviews (<=3)'],
        'Count': [pos_count, neg_count]
    })

    # Load Negative Reviews
    neg_df = pd.read_excel('Android_HearablesApp_Review.xlsx', sheet_name='Negative_Reviews')
    reason_counts = neg_df['Consolidated Reason'].value_counts()
    summary_df = pd.DataFrame({
        'Summary (-ve reviews)': reason_counts.index,
        'Count': reason_counts.values
    })

    # Helper: multiline list excluding Unknown Device
    def make_multiline(df_filter):
        df_filter = df_filter[df_filter['DUT'] != 'Unknown Device']
        if df_filter.empty:
            return "No valid devices found"
        counts = df_filter['DUT'].value_counts()
        return "\n".join([f"{dev} - {cnt}" for dev, cnt in counts.items()])

    # Remarks column
    summary_df['Remarks'] = [
        make_multiline(neg_df[neg_df['Consolidated Reason'] == reason])
        for reason in summary_df['Summary (-ve reviews)']
    ]

    if 'Positive Feedback' in summary_df['Summary (-ve reviews)'].values:
        summary_df.loc[
            summary_df['Summary (-ve reviews)'] == 'Positive Feedback', 'Remarks'
        ] = "Reviews are good but ratings are <= 3"

    # Workbook setup
    wb = load_workbook('Android_HearablesApp_Review.xlsx')
    ws = wb['Summary'] if 'Summary' in wb.sheetnames else wb.create_sheet('Summary')
    ws.delete_rows(1, ws.max_row)

    # Title
    ws.append(["📊 Review Summary Report"])
    ws.merge_cells('A1:C1')
    title = ws['A1']
    title.font = Font(size=14, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    ws.append([])

    # Write all data sections
    for df_part in [app_df, pd.DataFrame(), counts_df, pd.DataFrame(), summary_df]:
        for r in dataframe_to_rows(df_part, index=False):
            ws.append(r)

    # Basic styles
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical='top')
    blue_header = PatternFill(start_color="A7C7E7", end_color="A7C7E7", fill_type="solid")
    green_section = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

    # 📏 Increased Column C width
    col_widths = {'A': 25, 'B': 15, 'C': 35, 'D': 23}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Apply border + wrap
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = wrap

    # Highlight headers/sections
    for row in ws.iter_rows(min_row=1, max_col=3):
        values = [str(c.value).strip().lower() if c.value else "" for c in row]
        if "category" in values or "summary (-ve reviews)" in values:
            for cell in row:
                cell.fill = blue_header
                cell.font = Font(bold=True)
        elif "app name" in values:
            for cell in row:
                cell.fill = green_section
                cell.font = Font(bold=True)

    # 🧭 Center align Column C (Remarks)
    for row in ws.iter_rows():
        for cell in row:
            if hasattr(cell, "column_letter") and cell.column_letter == 'C':
                cell.alignment = Alignment(
                    wrap_text=True, horizontal='center', vertical='center'
                )

    # Dynamic row height for Column C
    def calc_row_height(text, col_width):
        if not isinstance(text, str) or not text.strip():
            return 15
        lines = text.split("\n")
        est_lines = sum(max(1, math.ceil(len(line) / (col_width * 1.5))) for line in lines)
        return min(300, 15 * est_lines)

    col_c_width = ws.column_dimensions['C'].width or 35
    for row in ws.iter_rows(min_row=1, max_col=3):
        for cell in row:
            if not hasattr(cell, "column_letter"):
                continue
            if cell.column_letter == 'C':
                ws.row_dimensions[cell.row].height = calc_row_height(cell.value, col_c_width)

    wb.save('Android_HearablesApp_Review.xlsx')
    print("✅ Summary sheet updated — Column C centered, wider, and formatted cleanly.")

def Summary_Enhance():
    # Load the workbook
    wb = load_workbook('Android_HearablesApp_Review.xlsx')

    # Select the "Summary" sheet
    sheet = wb['Summary']

    # Make the 1st row from column A to D bold and set background color to blue
    for col in range(1, 5):  # Columns A to D
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="B9D9EB", end_color="B9D9EB", fill_type="solid")
        #cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

    # Apply center alignment to columns B, C, and D
    for col in range(2, 5):  # Columns B to D
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Make the 4th row from column A to D bold and set background color
    for col in range(1, 3):  # Columns A to D
        cell = sheet.cell(row=4, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Make the 8th row from column A to D bold and set background color to yellow
    for col in range(1, 3):  # Columns A to D
        cell = sheet.cell(row=6, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Make the 8th row from column A to D bold and set background color
    for col in range(1, 3):  # Columns A to D (1 to 4 inclusive)
        cell = sheet.cell(row=5, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="66CC66", end_color="66CC66", fill_type="solid")


    # Make the 8th row from column A to D bold and set background color
    for col in range(1, 4):  # Columns A to D
        cell = sheet.cell(row=8, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Save the workbook (same file)
    wb.save('Android_HearablesApp_Review.xlsx')
    print("Summary Sheet's Last changes has been updated")

def close_excel_file(file_path):
    """
    Close the Excel file to release any locks on it.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        workbook.close()
        print("Closed the Excel file.")
    except Exception as e:
        print(f"Error closing the Excel file: {e}")


if __name__ == '__main__':
    HearablesApp_WebScraping()
    time.sleep(3)
    All_HtmlFiles_Csv()
    time.sleep(2)

    # IMPORTANT: Call App_Rating_Finder BEFORE Remove_Html()
    App_Rating_Finder()

    Csv_Date_Rating_Filter_7days()
    #Csv_Date_Rating_Filter_Alldate()
    CSV_Remove_Duplicates()
    Csv_2_Xlsx()
    SheetNameChange()
    Positive_Reviews()
    Negative_Reviews()
    Sorting_Xlsx('Positive_Reviews')
    Sorting_Xlsx('Negative_Reviews')
    Sorting_Xlsx('Happ_Review_Data')
    Negative_Review_Analysis()
    App_Version_Finder()
    Review_Date_Calculator()
    Summary()
    time.sleep(3)
    close_excel_file('Android_HearablesApp_Review.xlsx')

    # Remove HTML files LAST, after all processing is done
    Remove_Html()

    # Specify file names and remove CSV files
    file1 = 'Hearables_Data.csv'
    file2 = 'HearablesAppGCWeb.csv'

    # Remove the files after execution is completed
    try:
        os.remove(file1)
        print(f"{file1} has been deleted successfully.")
    except FileNotFoundError:
        print(f"{file1} not found.")

    try:
        os.remove(file2)
        print(f"{file2} has been deleted successfully.")
    except FileNotFoundError:
        print(f"{file2} not found.")

