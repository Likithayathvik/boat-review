# ============================================================
# main.py — boAt HApp Review Analyzer (Cloud Version)
# Author  : boAt R&D (QA)
# Replaces: HearablesApp_WebScraping, All_HtmlFiles_Csv,
#           App_Rating_Finder, Csv_Date_Rating_Filter_7days,
#           CSV_Remove_Duplicates, Csv_2_Xlsx
# ============================================================

import os
import uuid
import time
import openpyxl
import pandas as pd
import BFAT_HApp_PlayStore_Code as orig
from datetime import datetime, timedelta
from google_play_scraper import Sort, reviews

from BFAT_HApp_PlayStore_Code import (
    SheetNameChange,
    Positive_Reviews,
    Negative_Reviews,
    Sorting_Xlsx,
    Negative_Review_Analysis,
    Review_Date_Calculator,
    Summary,
    close_excel_file,
)

# ── CONFIG ──────────────────────────────────────────────────
APP_ID    = 'com.boAt.hearables'
XLSX_FILE = 'Android_HearablesApp_Review.xlsx'

# ── GLOBALS ─────────────────────────────────────────────────
App_Rating  = "N/A"   # not available from public Play Store
App_Version = "N/A"
lowest_date = None
latest_date = None

# ── STEP 1: FETCH REVIEWS ───────────────────────────────────
def fetch_and_save_reviews():
    print("=" * 60)
    print("Step 1: Fetching reviews from Google Play Store...")
    print("=" * 60)

    result, _ = reviews(
        APP_ID,
        lang='en',
        country='in',
        sort=Sort.NEWEST,
        count=600,  # 50/page x 12 pages — same as original Selenium
    )

    df = pd.DataFrame(result)

    # Filter last 7 days
    df['at'] = pd.to_datetime(df['at'])
    seven_days_ago = datetime.now() - timedelta(days=7)
    df = df[df['at'] >= seven_days_ago]

    # Add unique reviewId
    df['reviewId'] = [str(uuid.uuid4()) for _ in range(len(df))]

    # Rename columns
    df = df.rename(columns={
        'userName':             'User Name',
        'score':                'Ratings',
        'content':              'Reviews',
        'at':                   'Date Time',
        'reviewCreatedVersion': 'App Version',
        'replyContent':         'boAt Reply',
    })

    # Keep only useful columns
    df = df[[
        'reviewId', 'User Name', 'Ratings', 'Reviews',
        'Date Time', 'App Version', 'boAt Reply'
    ]]

    # Sort and deduplicate
    df = df.sort_values('Date Time', ascending=False)
    df = df.drop_duplicates(subset=['User Name', 'Ratings', 'Reviews'])

    # Save to Excel
    df.to_excel(XLSX_FILE, index=False)

    print(f"  Fetched : {len(df)} reviews from last 7 days")
    print(f"  Saved   : {XLSX_FILE}")

# ── STEP 2: GET APP VERSION ─────────────────────────────────
def App_Version_Finder():
    global App_Version

    workbook = openpyxl.load_workbook(XLSX_FILE)
    sheet    = workbook.active

    APP_VERSION_COL = 6   # column F — App Version in new layout
    highest_version = None

    for row in range(2, sheet.max_row + 1):
        ver = sheet.cell(row=row, column=APP_VERSION_COL).value
        if ver:
            ver_str = str(ver).split(":")[-1].strip()
            if highest_version is None or ver_str > highest_version:
                highest_version = ver_str

    App_Version = highest_version if highest_version else "N/A"
    print(f"  Highest App Version: {App_Version}")
    workbook.close()

# ── MAIN ────────────────────────────────────────────────────
if __name__ == '__main__':

    # Step 1 — Fetch reviews from Play Store
    fetch_and_save_reviews()
    time.sleep(2)

    # Step 2 — Rename Sheet1 → Happ_Review_Data
    SheetNameChange()

    # Step 3 — Split into Positive / Negative sheets
    Positive_Reviews()
    Negative_Reviews()

    # Step 4 — Format all sheets
    Sorting_Xlsx('Positive_Reviews')
    Sorting_Xlsx('Negative_Reviews')
    Sorting_Xlsx('Happ_Review_Data')

    # Step 5 — Analysis
    Negative_Review_Analysis()

    # Step 6 — Get App Version + Date Range
    App_Version_Finder()
    Review_Date_Calculator()

    # Step 7 — Inject globals into original script before Summary
    orig.App_Version = App_Version
    orig.App_Rating  = App_Rating
    orig.lowest_date = lowest_date
    orig.latest_date = latest_date

    # Step 8 — Generate Summary sheet
    Summary()
    time.sleep(3)

    # Step 9 — Close file
    close_excel_file(XLSX_FILE)

    print("=" * 60)
    print("Pipeline complete! Android_HearablesApp_Review.xlsx ready.")
    print("=" * 60)